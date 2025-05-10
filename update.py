import os
import json
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from qualer_sdk import Configuration, ApiClient, AssetServiceRecordsApi

# ---- Config ----
FILE_PATH = r"https://jgiquality.sharepoint.com/sites/JGI/Shared%20Documents/Pyro/WireSetDisambiguation.xlsx"  # noqa: E501
TABLE_SHEET = "WireSets"
TABLE_NAME = "Table1"
TOKEN = os.environ.get("QUALER_API_KEY")

# ---- Qualer API Setup ----
config = Configuration()
config.host = "https://jgiquality.qualer.com"
client = ApiClient(configuration=config)
client.default_headers["Authorization"] = f"Api-Token {TOKEN}"
service_api = AssetServiceRecordsApi(client)

# ---- Load Excel and Table1 ----
excel = pd.read_excel(FILE_PATH, sheet_name=TABLE_SHEET, engine="openpyxl")
if "AssetId" not in excel.columns:
    raise ValueError("Expected 'AssetId' column in Table1")

# ---- Pull latest service record per asset ----
records = []
for _, row in tqdm(excel.iterrows(), total=len(excel)):
    asset_id = row["AssetId"]
    try:
        service_records = service_api.asset_service_records_get_asset_service_records_by_asset(asset_id)  # noqa: E501
        if not service_records:
            latest = None
        else:
            latest = max(service_records, key=lambda r: r.service_date or "")
        record_json = json.dumps(latest.to_dict() if latest else {}, indent=2)
    except Exception as e:
        record_json = json.dumps({"error": str(e)})
    records.append(record_json)

# ---- Write back into DataFrame ----
excel["LatestServiceRecordRaw"] = records

# ---- Save back into Excel (overwrite) ----
wb = load_workbook(FILE_PATH)
ws = wb[TABLE_SHEET]

# Remove the old table rows
start_row = 2  # Assuming headers start at row 1
ws.delete_rows(start_row, ws.max_row - 1)

# Write updated rows
updated_rows = dataframe_to_rows(excel, index=False, header=False)
for r_idx, row in enumerate(updated_rows, start=start_row):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(FILE_PATH)
print("✅ WireSetDisambiguation.xlsx updated.")
